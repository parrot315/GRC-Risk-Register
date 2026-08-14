"""
app.py
------
Main Flask application for the GRC Risk Register academic project.

Title:  Risk Register Development Based on ISO/IEC 27001 for a
        Simulated Mid-Sized Organization
Org:    ParrotSec Technologies (fictional, IT / Software Services, ~250 employees)

Run locally with:
    python app.py

This is an academic / simulated GRC application. It is NOT an officially
certified ISO/IEC 27001 compliance tool.
"""

import os
import io
import sqlite3
from datetime import datetime

from flask import (
    Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
)

from database_setup import init_db, get_connection, calculate_level

# ---------------------------------------------------------------------------
# App configuration
# ---------------------------------------------------------------------------
app = Flask(__name__)
# Secret key is generated locally at runtime and only used for flash messages
# / session signing on this local academic instance. Do not reuse in production.
app.secret_key = os.environ.get("GRC_SECRET_KEY", os.urandom(24).hex())

EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

LIKELIHOOD_LABELS = {1: "Rare", 2: "Unlikely", 3: "Possible", 4: "Likely", 5: "Almost Certain"}
IMPACT_LABELS = {1: "Insignificant", 2: "Minor", 3: "Moderate", 4: "Major", 5: "Severe"}
TREATMENT_OPTIONS = ["Mitigate", "Avoid", "Transfer", "Accept"]
STATUS_OPTIONS = ["Open", "In Progress", "Treated", "Accepted", "Closed"]

LEVEL_ORDER = {"Low": 1, "Medium": 2, "High": 3, "Critical": 4}


# ---------------------------------------------------------------------------
# Template filters / global context
# ---------------------------------------------------------------------------
@app.context_processor
def inject_globals():
    return {
        "likelihood_labels": LIKELIHOOD_LABELS,
        "impact_labels": IMPACT_LABELS,
        "treatment_options": TREATMENT_OPTIONS,
        "status_options": STATUS_OPTIONS,
        "current_year": datetime.now().year,
    }


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------
def validate_risk_form(form):
    """Validate submitted risk form data. Returns (errors list, cleaned dict)."""
    errors = []
    cleaned = {}

    def get_int(field, label):
        val = form.get(field, "").strip()
        if val == "":
            errors.append(f"{label} is required.")
            return None
        try:
            ival = int(val)
        except ValueError:
            errors.append(f"{label} must be a whole number.")
            return None
        if not (1 <= ival <= 5):
            errors.append(f"{label} must be between 1 and 5.")
            return None
        return ival

    required_text_fields = {
        "threat": "Threat",
        "vulnerability": "Vulnerability",
        "risk_treatment": "Risk treatment",
    }
    for field, label in required_text_fields.items():
        val = (form.get(field) or "").strip()
        if not val:
            errors.append(f"{label} is required.")
        cleaned[field] = val

    if cleaned.get("risk_treatment") and cleaned["risk_treatment"] not in TREATMENT_OPTIONS:
        errors.append("Risk treatment must be one of: " + ", ".join(TREATMENT_OPTIONS))

    cleaned["likelihood"] = get_int("likelihood", "Likelihood")
    cleaned["impact"] = get_int("impact", "Impact")

    # Residual likelihood/impact are optional but validated if provided
    res_like_raw = (form.get("residual_likelihood") or "").strip()
    res_impact_raw = (form.get("residual_impact") or "").strip()
    cleaned["residual_likelihood"] = None
    cleaned["residual_impact"] = None
    if res_like_raw:
        try:
            rl = int(res_like_raw)
            if not (1 <= rl <= 5):
                errors.append("Residual likelihood must be between 1 and 5.")
            else:
                cleaned["residual_likelihood"] = rl
        except ValueError:
            errors.append("Residual likelihood must be a whole number.")
    if res_impact_raw:
        try:
            ri = int(res_impact_raw)
            if not (1 <= ri <= 5):
                errors.append("Residual impact must be between 1 and 5.")
            else:
                cleaned["residual_impact"] = ri
        except ValueError:
            errors.append("Residual impact must be a whole number.")

    status = (form.get("status") or "Open").strip()
    if status not in STATUS_OPTIONS:
        errors.append("Invalid status value.")
    cleaned["status"] = status

    cleaned["asset_id"] = form.get("asset_id") or None
    cleaned["asset_owner"] = (form.get("asset_owner") or "").strip()
    cleaned["risk_description"] = (form.get("risk_description") or "").strip()
    cleaned["existing_controls"] = (form.get("existing_controls") or "").strip()
    cleaned["treatment_description"] = (form.get("treatment_description") or "").strip()
    cleaned["proposed_controls"] = (form.get("proposed_controls") or "").strip()
    cleaned["risk_owner"] = (form.get("risk_owner") or "").strip()
    cleaned["target_date"] = (form.get("target_date") or "").strip()
    cleaned["notes"] = (form.get("notes") or "").strip()
    cleaned["iso_controls"] = form.getlist("iso_controls")

    return errors, cleaned


def next_risk_code(conn):
    """Generate the next sequential risk code, e.g. R-018."""
    row = conn.execute("SELECT risk_code FROM risks ORDER BY id DESC LIMIT 1").fetchone()
    if not row:
        return "R-001"
    try:
        last_num = int(row["risk_code"].split("-")[1])
    except (IndexError, ValueError):
        last_num = conn.execute("SELECT COUNT(*) as c FROM risks").fetchone()["c"]
    return f"R-{last_num + 1:03d}"


def row_to_dict(row):
    return dict(row) if row else None


# ---------------------------------------------------------------------------
# ROUTES: Dashboard
# ---------------------------------------------------------------------------
@app.route("/")
def dashboard():
    conn = get_connection()

    total_risks = conn.execute("SELECT COUNT(*) c FROM risks").fetchone()["c"]

    def count_level(field):
        rows = conn.execute(f"SELECT {field} AS lvl, COUNT(*) c FROM risks GROUP BY {field}").fetchall()
        return {r["lvl"]: r["c"] for r in rows}

    inherent_counts = count_level("inherent_level")
    critical = inherent_counts.get("Critical", 0)
    high = inherent_counts.get("High", 0)
    medium = inherent_counts.get("Medium", 0)
    low = inherent_counts.get("Low", 0)

    status_counts_rows = conn.execute(
        "SELECT status, COUNT(*) c FROM risks GROUP BY status"
    ).fetchall()
    status_counts = {r["status"]: r["c"] for r in status_counts_rows}
    open_risks = status_counts.get("Open", 0) + status_counts.get("In Progress", 0)
    treated_risks = status_counts.get("Treated", 0) + status_counts.get("Closed", 0)
    accepted_risks = status_counts.get("Accepted", 0)

    top_risks = conn.execute("""
        SELECT r.*, a.name as asset_name
        FROM risks r
        LEFT JOIN assets a ON r.asset_id = a.id
        ORDER BY r.inherent_score DESC
        LIMIT 5
    """).fetchall()

    recent_risks = conn.execute("""
        SELECT r.*, a.name as asset_name
        FROM risks r
        LEFT JOIN assets a ON r.asset_id = a.id
        ORDER BY r.created_at DESC, r.id DESC
        LIMIT 5
    """).fetchall()

    # Matrix data: counts per (likelihood, impact) cell for the risk matrix widget
    matrix_rows = conn.execute("""
        SELECT likelihood, impact, COUNT(*) c FROM risks GROUP BY likelihood, impact
    """).fetchall()
    matrix_counts = {(r["likelihood"], r["impact"]): r["c"] for r in matrix_rows}

    organization = conn.execute("SELECT * FROM organization WHERE id = 1").fetchone()

    conn.close()

    return render_template(
        "dashboard.html",
        organization=organization,
        total_risks=total_risks,
        critical=critical, high=high, medium=medium, low=low,
        open_risks=open_risks, treated_risks=treated_risks, accepted_risks=accepted_risks,
        top_risks=top_risks, recent_risks=recent_risks,
        matrix_counts=matrix_counts,
        level_order=LEVEL_ORDER,
    )


# ---------------------------------------------------------------------------
# ROUTES: Risk Register (list, search, filter, sort)
# ---------------------------------------------------------------------------
@app.route("/risks")
def risks_list():
    conn = get_connection()

    search = request.args.get("q", "").strip()
    level_filter = request.args.get("level", "").strip()
    status_filter = request.args.get("status", "").strip()
    treatment_filter = request.args.get("treatment", "").strip()
    sort = request.args.get("sort", "inherent_score")
    direction = request.args.get("dir", "desc")

    allowed_sort_columns = {
        "risk_code", "threat", "vulnerability", "likelihood", "impact",
        "inherent_score", "inherent_level", "risk_treatment", "status",
        "residual_score", "target_date", "created_at"
    }
    if sort not in allowed_sort_columns:
        sort = "inherent_score"
    direction = "ASC" if direction.lower() == "asc" else "DESC"

    query = """
        SELECT r.*, a.name as asset_name
        FROM risks r
        LEFT JOIN assets a ON r.asset_id = a.id
        WHERE 1=1
    """
    params = []

    if search:
        query += """ AND (
            r.risk_code LIKE ? OR r.threat LIKE ? OR r.vulnerability LIKE ?
            OR r.risk_description LIKE ? OR a.name LIKE ? OR r.risk_owner LIKE ?
        )"""
        like = f"%{search}%"
        params.extend([like, like, like, like, like, like])

    if level_filter:
        query += " AND r.inherent_level = ?"
        params.append(level_filter)

    if status_filter:
        query += " AND r.status = ?"
        params.append(status_filter)

    if treatment_filter:
        query += " AND r.risk_treatment = ?"
        params.append(treatment_filter)

    query += f" ORDER BY r.{sort} {direction}"

    risks = conn.execute(query, params).fetchall()
    conn.close()

    return render_template(
        "risks.html",
        risks=risks,
        search=search,
        level_filter=level_filter,
        status_filter=status_filter,
        treatment_filter=treatment_filter,
        sort=sort,
        direction=direction.lower(),
    )


# ---------------------------------------------------------------------------
# ROUTES: Add Risk
# ---------------------------------------------------------------------------
@app.route("/risks/add", methods=["GET", "POST"])
def add_risk():
    conn = get_connection()
    assets = conn.execute("SELECT * FROM assets ORDER BY name").fetchall()
    iso_controls = conn.execute("SELECT * FROM iso_controls ORDER BY control_id").fetchall()

    if request.method == "POST":
        errors, data = validate_risk_form(request.form)

        if errors:
            for e in errors:
                flash(e, "danger")
            conn.close()
            return render_template(
                "add_risk.html", assets=assets, iso_controls=iso_controls, form=request.form
            )

        inherent_score = data["likelihood"] * data["impact"]
        inherent_level = calculate_level(inherent_score)

        residual_score = None
        residual_level = None
        if data["residual_likelihood"] and data["residual_impact"]:
            residual_score = data["residual_likelihood"] * data["residual_impact"]
            residual_level = calculate_level(residual_score)

        risk_code = next_risk_code(conn)

        cur = conn.execute("""
            INSERT INTO risks (
                risk_code, asset_id, asset_owner, threat, vulnerability, risk_description,
                existing_controls, likelihood, impact, inherent_score, inherent_level,
                risk_treatment, treatment_description, proposed_controls, risk_owner,
                target_date, status, residual_likelihood, residual_impact,
                residual_score, residual_level, notes, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (
            risk_code, data["asset_id"], data["asset_owner"], data["threat"], data["vulnerability"],
            data["risk_description"], data["existing_controls"], data["likelihood"], data["impact"],
            inherent_score, inherent_level, data["risk_treatment"], data["treatment_description"],
            data["proposed_controls"], data["risk_owner"], data["target_date"], data["status"],
            data["residual_likelihood"], data["residual_impact"], residual_score, residual_level,
            data["notes"]
        ))
        new_id = cur.lastrowid

        for ctrl_id in data["iso_controls"]:
            conn.execute(
                "INSERT OR IGNORE INTO risk_controls (risk_id, control_id) VALUES (?, ?)",
                (new_id, ctrl_id)
            )

        conn.commit()
        conn.close()
        flash(f"Risk {risk_code} added successfully.", "success")
        return redirect(url_for("risk_details", risk_id=new_id))

    conn.close()
    return render_template("add_risk.html", assets=assets, iso_controls=iso_controls, form=request.form)


# ---------------------------------------------------------------------------
# ROUTES: Edit Risk
# ---------------------------------------------------------------------------
@app.route("/risks/edit/<int:risk_id>", methods=["GET", "POST"])
def edit_risk(risk_id):
    conn = get_connection()
    risk = conn.execute("SELECT * FROM risks WHERE id = ?", (risk_id,)).fetchone()
    if not risk:
        conn.close()
        flash("Risk not found.", "danger")
        return redirect(url_for("risks_list"))

    assets = conn.execute("SELECT * FROM assets ORDER BY name").fetchall()
    iso_controls = conn.execute("SELECT * FROM iso_controls ORDER BY control_id").fetchall()
    selected_controls = {
        r["control_id"] for r in conn.execute(
            "SELECT control_id FROM risk_controls WHERE risk_id = ?", (risk_id,)
        ).fetchall()
    }

    if request.method == "POST":
        errors, data = validate_risk_form(request.form)

        if errors:
            for e in errors:
                flash(e, "danger")
            conn.close()
            merged = dict(risk)
            merged.update(request.form)
            return render_template(
                "edit_risk.html", risk=merged, assets=assets, iso_controls=iso_controls,
                selected_controls=set(request.form.getlist("iso_controls")), risk_id=risk_id
            )

        inherent_score = data["likelihood"] * data["impact"]
        inherent_level = calculate_level(inherent_score)

        residual_score = None
        residual_level = None
        if data["residual_likelihood"] and data["residual_impact"]:
            residual_score = data["residual_likelihood"] * data["residual_impact"]
            residual_level = calculate_level(residual_score)

        conn.execute("""
            UPDATE risks SET
                asset_id = ?, asset_owner = ?, threat = ?, vulnerability = ?, risk_description = ?,
                existing_controls = ?, likelihood = ?, impact = ?, inherent_score = ?, inherent_level = ?,
                risk_treatment = ?, treatment_description = ?, proposed_controls = ?, risk_owner = ?,
                target_date = ?, status = ?, residual_likelihood = ?, residual_impact = ?,
                residual_score = ?, residual_level = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (
            data["asset_id"], data["asset_owner"], data["threat"], data["vulnerability"],
            data["risk_description"], data["existing_controls"], data["likelihood"], data["impact"],
            inherent_score, inherent_level, data["risk_treatment"], data["treatment_description"],
            data["proposed_controls"], data["risk_owner"], data["target_date"], data["status"],
            data["residual_likelihood"], data["residual_impact"], residual_score, residual_level,
            data["notes"], risk_id
        ))

        conn.execute("DELETE FROM risk_controls WHERE risk_id = ?", (risk_id,))
        for ctrl_id in data["iso_controls"]:
            conn.execute(
                "INSERT OR IGNORE INTO risk_controls (risk_id, control_id) VALUES (?, ?)",
                (risk_id, ctrl_id)
            )

        conn.commit()
        conn.close()
        flash(f"Risk {risk['risk_code']} updated successfully.", "success")
        return redirect(url_for("risk_details", risk_id=risk_id))

    conn.close()
    return render_template(
        "edit_risk.html", risk=risk, assets=assets, iso_controls=iso_controls,
        selected_controls=selected_controls, risk_id=risk_id
    )


# ---------------------------------------------------------------------------
# ROUTES: Delete Risk
# ---------------------------------------------------------------------------
@app.route("/risks/delete/<int:risk_id>", methods=["POST"])
def delete_risk(risk_id):
    conn = get_connection()
    risk = conn.execute("SELECT risk_code FROM risks WHERE id = ?", (risk_id,)).fetchone()
    if risk:
        conn.execute("DELETE FROM risk_controls WHERE risk_id = ?", (risk_id,))
        conn.execute("DELETE FROM risks WHERE id = ?", (risk_id,))
        conn.commit()
        flash(f"Risk {risk['risk_code']} deleted.", "success")
    else:
        flash("Risk not found.", "danger")
    conn.close()
    return redirect(url_for("risks_list"))


# ---------------------------------------------------------------------------
# ROUTES: Risk Details
# ---------------------------------------------------------------------------
@app.route("/risks/<int:risk_id>")
def risk_details(risk_id):
    conn = get_connection()
    risk = conn.execute("""
        SELECT r.*, a.name as asset_name, a.category as asset_category
        FROM risks r
        LEFT JOIN assets a ON r.asset_id = a.id
        WHERE r.id = ?
    """, (risk_id,)).fetchone()

    if not risk:
        conn.close()
        flash("Risk not found.", "danger")
        return redirect(url_for("risks_list"))

    mapped_controls = conn.execute("""
        SELECT ic.* FROM iso_controls ic
        JOIN risk_controls rc ON rc.control_id = ic.id
        WHERE rc.risk_id = ?
        ORDER BY ic.control_id
    """, (risk_id,)).fetchall()

    conn.close()
    return render_template("risk_details.html", risk=risk, mapped_controls=mapped_controls)


# ---------------------------------------------------------------------------
# ROUTES: Risk Matrix
# ---------------------------------------------------------------------------
@app.route("/matrix")
def matrix():
    conn = get_connection()
    rows = conn.execute("""
        SELECT id, risk_code, threat, likelihood, impact, inherent_level
        FROM risks
    """).fetchall()
    conn.close()

    # Build a 5x5 grid: grid[impact][likelihood] = list of risks
    grid = {i: {l: [] for l in range(1, 6)} for i in range(1, 6)}
    for r in rows:
        grid[r["impact"]][r["likelihood"]].append(r)

    return render_template("matrix.html", grid=grid)


# ---------------------------------------------------------------------------
# ROUTES: ISO 27001 Controls
# ---------------------------------------------------------------------------
@app.route("/controls")
def controls():
    conn = get_connection()
    controls_rows = conn.execute("SELECT * FROM iso_controls ORDER BY control_id").fetchall()

    # For each control, count how many risks are mapped to it
    mapping_counts = {
        r["control_id"]: r["c"] for r in conn.execute("""
            SELECT control_id, COUNT(*) c FROM risk_controls GROUP BY control_id
        """).fetchall()
    }
    conn.close()

    themes = {}
    for c in controls_rows:
        themes.setdefault(c["theme"], []).append(c)

    return render_template("controls.html", themes=themes, mapping_counts=mapping_counts)


@app.route("/controls/<int:control_id>")
def control_risks(control_id):
    conn = get_connection()
    control = conn.execute("SELECT * FROM iso_controls WHERE id = ?", (control_id,)).fetchone()
    mapped_risks = conn.execute("""
        SELECT r.* FROM risks r
        JOIN risk_controls rc ON rc.risk_id = r.id
        WHERE rc.control_id = ?
        ORDER BY r.inherent_score DESC
    """, (control_id,)).fetchall()
    conn.close()
    if not control:
        flash("Control not found.", "danger")
        return redirect(url_for("controls"))
    return render_template("control_risks.html", control=control, mapped_risks=mapped_risks)


# ---------------------------------------------------------------------------
# ROUTES: Reports / Exports
# ---------------------------------------------------------------------------
@app.route("/reports")
def reports():
    conn = get_connection()
    total_risks = conn.execute("SELECT COUNT(*) c FROM risks").fetchone()["c"]
    conn.close()
    return render_template("reports.html", total_risks=total_risks,
                            generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"))


@app.route("/reports/export/excel/<report_type>")
def export_excel(report_type):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_connection()

    org_row = conn.execute("SELECT name FROM organization WHERE id = 1").fetchone()
    org_name_slug = (org_row["name"] if org_row else "Organization").replace(" ", "_")

    wb = openpyxl.Workbook()
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(*(Side(style="thin"),) * 4)

    def style_header(row_idx, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    if report_type == "register":
        ws.title = "Risk Register"
        headers = [
            "Risk ID", "Asset", "Asset Owner", "Threat", "Vulnerability", "Risk Description",
            "Existing Controls", "Likelihood", "Impact", "Inherent Score", "Inherent Level",
            "Treatment", "Treatment Description", "Proposed Controls", "Risk Owner",
            "Target Date", "Status", "Residual Likelihood", "Residual Impact",
            "Residual Score", "Residual Level", "Notes"
        ]
        ws.append(headers)
        style_header(1, len(headers))

        rows = conn.execute("""
            SELECT r.*, a.name as asset_name FROM risks r
            LEFT JOIN assets a ON r.asset_id = a.id
            ORDER BY r.inherent_score DESC
        """).fetchall()
        for r in rows:
            ws.append([
                r["risk_code"], r["asset_name"], r["asset_owner"], r["threat"], r["vulnerability"],
                r["risk_description"], r["existing_controls"], r["likelihood"], r["impact"],
                r["inherent_score"], r["inherent_level"], r["risk_treatment"], r["treatment_description"],
                r["proposed_controls"], r["risk_owner"], r["target_date"], r["status"],
                r["residual_likelihood"], r["residual_impact"], r["residual_score"], r["residual_level"],
                r["notes"]
            ])
        filename = f"Risk_Register_{org_name_slug}.xlsx"

    elif report_type == "summary":
        ws.title = "Risk Summary"
        headers = ["Metric", "Value"]
        ws.append(headers)
        style_header(1, len(headers))

        total = conn.execute("SELECT COUNT(*) c FROM risks").fetchone()["c"]
        level_rows = conn.execute(
            "SELECT inherent_level, COUNT(*) c FROM risks GROUP BY inherent_level"
        ).fetchall()
        level_counts = {r["inherent_level"]: r["c"] for r in level_rows}
        status_rows = conn.execute("SELECT status, COUNT(*) c FROM risks GROUP BY status").fetchall()
        status_counts = {r["status"]: r["c"] for r in status_rows}
        treatment_rows = conn.execute(
            "SELECT risk_treatment, COUNT(*) c FROM risks GROUP BY risk_treatment"
        ).fetchall()
        treatment_counts = {r["risk_treatment"]: r["c"] for r in treatment_rows}

        summary_data = [
            ("Total Risks", total),
            ("Critical Risks", level_counts.get("Critical", 0)),
            ("High Risks", level_counts.get("High", 0)),
            ("Medium Risks", level_counts.get("Medium", 0)),
            ("Low Risks", level_counts.get("Low", 0)),
        ]
        for status in STATUS_OPTIONS:
            summary_data.append((f"Status: {status}", status_counts.get(status, 0)))
        for treatment in TREATMENT_OPTIONS:
            summary_data.append((f"Treatment: {treatment}", treatment_counts.get(treatment, 0)))

        for row in summary_data:
            ws.append(row)
        filename = f"Risk_Summary_{org_name_slug}.xlsx"

    elif report_type == "iso_mapping":
        ws.title = "ISO 27001 Mapping"
        headers = ["Control ID", "Control Title", "Theme", "Mapped Risks (Count)", "Mapped Risk IDs"]
        ws.append(headers)
        style_header(1, len(headers))

        controls_rows = conn.execute("SELECT * FROM iso_controls ORDER BY control_id").fetchall()
        for c in controls_rows:
            mapped = conn.execute("""
                SELECT r.risk_code FROM risks r
                JOIN risk_controls rc ON rc.risk_id = r.id
                WHERE rc.control_id = ?
                ORDER BY r.risk_code
            """, (c["id"],)).fetchall()
            risk_codes = ", ".join([m["risk_code"] for m in mapped])
            ws.append([c["control_id"], c["title"], c["theme"], len(mapped), risk_codes])
        filename = f"ISO27001_Control_Mapping_{org_name_slug}.xlsx"

    else:
        conn.close()
        flash("Unknown report type requested.", "danger")
        return redirect(url_for("reports"))

    conn.close()

    # Auto-fit column widths (approximate)
    for col_cells in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    ws.freeze_panes = "A2"

    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route("/reports/print/<report_type>")
def print_report(report_type):
    """Printable HTML report the user can save as PDF via the browser's print dialog."""
    conn = get_connection()

    if report_type == "register":
        rows = conn.execute("""
            SELECT r.*, a.name as asset_name FROM risks r
            LEFT JOIN assets a ON r.asset_id = a.id
            ORDER BY r.inherent_score DESC
        """).fetchall()
        conn.close()
        return render_template("print_register.html", risks=rows,
                                generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"))

    elif report_type == "iso_mapping":
        controls_rows = conn.execute("SELECT * FROM iso_controls ORDER BY control_id").fetchall()
        mapping = {}
        for c in controls_rows:
            mapped = conn.execute("""
                SELECT r.risk_code, r.threat FROM risks r
                JOIN risk_controls rc ON rc.risk_id = r.id
                WHERE rc.control_id = ?
                ORDER BY r.risk_code
            """, (c["id"],)).fetchall()
            mapping[c["id"]] = mapped
        conn.close()
        return render_template("print_iso_mapping.html", controls=controls_rows, mapping=mapping,
                                generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"))

    conn.close()
    flash("Unknown report type requested.", "danger")
    return redirect(url_for("reports"))


# ---------------------------------------------------------------------------
# ROUTES: About
# ---------------------------------------------------------------------------
@app.route("/about")
def about():
    conn = get_connection()
    organization = conn.execute("SELECT * FROM organization WHERE id = 1").fetchone()
    conn.close()
    return render_template("about.html", organization=organization)


# ---------------------------------------------------------------------------
# API endpoint used by JavaScript to preview calculated risk score live
# ---------------------------------------------------------------------------
@app.route("/api/calculate", methods=["POST"])
def api_calculate():
    data = request.get_json(silent=True) or {}
    try:
        likelihood = int(data.get("likelihood"))
        impact = int(data.get("impact"))
        if not (1 <= likelihood <= 5 and 1 <= impact <= 5):
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "likelihood and impact must be integers 1-5"}), 400

    score = likelihood * impact
    return jsonify({"score": score, "level": calculate_level(score)})


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------
@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404


@app.errorhandler(500)
def server_error(e):
    return render_template("500.html"), 500


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="127.0.0.1", port=5000)
